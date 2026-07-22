from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.core.mail import EmailMultiAlternatives
from django.utils import timezone
from accounts.mixins import BranchPermissionMixin, filter_by_branch
from django.contrib.messages.views import SuccessMessageMixin
from django.db.models import Q
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST, require_http_methods

from core.models import Branch
from courses.models import Course
from .models import StudentOffer, OfferRecipient, OfferNote
from .forms import StudentOfferForm, OfferRecipientForm, OfferRecipientAddForm, OfferNoteForm, RootQuickOfferForm
from .whatsapp import send_whatsapp_message, send_whatsapp_pdf, _is_root_branch
from prospects.models import Prospect


import logging

logger = logging.getLogger(__name__)


def _prepare_arabic(text):
    """Reshape and apply BiDi algorithm for Arabic text in ReportLab PDF.
    Escapes XML special chars so ReportLab Paragraph doesn't break."""
    if not text:
        return ''
    try:
        import arabic_reshaper
        from bidi.algorithm import get_display
        reshaped = arabic_reshaper.reshape(str(text))
        bidi_text = get_display(reshaped, base_dir='R')
        return bidi_text.replace('&', '&amp;')
    except Exception as exc:
        logger.warning('Arabic text preparation failed: %s', exc, exc_info=True)
        return str(text).replace('&', '&amp;')


def _prepare_arabic_paragraph(text):
    """Prepare multi-line Arabic text for ReportLab.
    Each line is BiDi-processed separately so bullet order is preserved."""
    if not text:
        return ''
    lines = str(text).splitlines()
    out_lines = []
    for line in lines:
        line = line.strip()
        if not line:
            continue
        bullet = ''
        if line.startswith('•') or line.startswith('-') or line.startswith('*'):
            # Find the bullet separator (first whitespace after bullet)
            parts = line.split(None, 1)
            if len(parts) == 2:
                bullet = parts[0]
                line = parts[1]
            else:
                bullet = ''
        prepared = _prepare_arabic(line)
        if bullet:
            # Append bullet at the end so it appears on the right in LTR render.
            prepared = f'{prepared} {bullet}'
        out_lines.append(prepared)
    return '<br/>'.join(out_lines)


def build_offer_pdf(offer, recipient=None):
    """Build offer PDF as BytesIO. Used by export and email attachment."""
    import os
    from io import BytesIO

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.units import cm
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from django.contrib.staticfiles.finders import find

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        rightMargin=1.5*cm, leftMargin=1.5*cm,
        topMargin=1.0*cm, bottomMargin=1.0*cm
    )
    elements = []
    styles = getSampleStyleSheet()

    # Register Arabic font
    from reportlab.pdfbase.pdfmetrics import registerFontFamily
    from django.conf import settings
    font_name = 'ArialArabic'
    BASE_DIR = settings.BASE_DIR

    candidates = [
        r'C:\Windows\Fonts\arial.ttf',
        os.path.join(BASE_DIR, 'static', 'fonts', 'Amiri-Regular.ttf'),
        os.path.join(BASE_DIR, 'staticfiles', 'fonts', 'Amiri-Regular.ttf'),
        os.path.join(BASE_DIR, 'static', 'fonts', 'Cairo-Regular.ttf'),
        os.path.join(BASE_DIR, 'staticfiles', 'fonts', 'Cairo-Regular.ttf'),
        find('fonts/Amiri-Regular.ttf'),
        find('fonts/Cairo-Regular.ttf'),
    ]

    registered = False
    for font_path in candidates:
        if font_path and os.path.exists(font_path):
            try:
                pdfmetrics.registerFont(TTFont(font_name, font_path))
                registerFontFamily(font_name, normal=font_name, bold=font_name, italic=font_name, boldItalic=font_name)
                registered = True
                break
            except Exception:
                continue

    if not registered:
        font_name = 'Helvetica'

    # Palette
    PRIMARY = colors.HexColor('#1e40af')
    PRIMARY_LIGHT = colors.HexColor('#dbeafe')
    SECONDARY = colors.HexColor('#475569')
    MUTED = colors.HexColor('#94a3b8')
    LIGHT_BG = colors.HexColor('#f8fafc')
    BORDER = colors.HexColor('#e2e8f0')
    WHITE = colors.HexColor('#ffffff')

    # Styles
    header_small_style = ParagraphStyle(
        'HeaderSmall', parent=styles['Normal'], fontName=font_name, fontSize=8, leading=10,
        alignment=TA_CENTER, textColor=SECONDARY, spaceAfter=1
    )
    title_style = ParagraphStyle(
        'DocTitle', parent=styles['Title'], fontName=font_name, fontSize=22, leading=28,
        alignment=TA_CENTER, textColor=WHITE, spaceAfter=0
    )
    section_label_style = ParagraphStyle(
        'SectionLabel', parent=styles['Normal'], fontName=font_name, fontSize=9, leading=12,
        alignment=TA_RIGHT, textColor=SECONDARY, spaceAfter=2
    )
    field_style = ParagraphStyle(
        'Field', parent=styles['Normal'], fontName=font_name, fontSize=11, leading=15,
        alignment=TA_RIGHT, textColor=colors.HexColor('#0f172a')
    )
    price_value_style = ParagraphStyle(
        'PriceValue', parent=styles['Normal'], fontName=font_name, fontSize=20, leading=26,
        alignment=TA_CENTER, textColor=PRIMARY
    )
    price_label_center_style = ParagraphStyle(
        'PriceLabelCenter', parent=styles['Normal'], fontName=font_name, fontSize=10, leading=14,
        alignment=TA_CENTER, textColor=PRIMARY
    )
    price_note_style = ParagraphStyle(
        'PriceNote', parent=styles['Normal'], fontName=font_name, fontSize=9, leading=12,
        alignment=TA_CENTER, textColor=SECONDARY
    )
    content_style = ParagraphStyle(
        'Content', parent=styles['Normal'], fontName=font_name, fontSize=10, leading=16,
        alignment=TA_RIGHT, textColor=colors.HexColor('#0f172a')
    )
    sig_style = ParagraphStyle(
        'Sig', parent=styles['Normal'], fontName=font_name, fontSize=10, leading=14,
        alignment=TA_CENTER, textColor=SECONDARY
    )
    footer_style = ParagraphStyle(
        'Footer', parent=styles['Normal'], fontName=font_name, fontSize=8, leading=11,
        alignment=TA_CENTER, textColor=MUTED
    )

    branch = offer.branch
    company = branch.company if branch else None

    # ========== HEADER: COMPANY LOGO + BRANCH LOGO + INFO ==========
    def _load_logo(image_field, width=2.0*cm, height=2.0*cm):
        if not image_field:
            return None
        try:
            return Image(image_field.path, width=width, height=height)
        except Exception:
            return None

    company_logo = _load_logo(company.logo if company else None)
    branch_logo = _load_logo(branch.logo if branch else None)

    header_lines = []
    if company:
        header_lines.append(_prepare_arabic(company.name))
    if branch:
        header_lines.append(_prepare_arabic(branch.name))
    contact_bits = []
    if branch:
        if branch.address:
            contact_bits.append(_prepare_arabic(branch.address))
        phones = [p for p in [branch.mobile, branch.phone1] if p]
        if phones:
            contact_bits.append(_prepare_arabic(' / '.join(phones)))
        if branch.email:
            contact_bits.append(_prepare_arabic(branch.email))
    if contact_bits:
        header_lines.append(' · '.join(contact_bits))

    header_text = '<br/>'.join(header_lines)
    header_para = Paragraph(header_text, header_small_style) if header_text else None

    if company_logo and branch_logo and header_para:
        header_inner = Table(
            [[company_logo, header_para, branch_logo]],
            colWidths=[doc.width*0.16, doc.width*0.68, doc.width*0.16]
        )
    elif company_logo and branch_logo:
        header_inner = Table(
            [[company_logo, '', branch_logo]],
            colWidths=[doc.width*0.16, doc.width*0.68, doc.width*0.16]
        )
    elif branch_logo and header_para:
        header_inner = Table(
            [[branch_logo, header_para]],
            colWidths=[doc.width*0.16, doc.width*0.84]
        )
    elif company_logo and header_para:
        header_inner = Table(
            [[company_logo, header_para]],
            colWidths=[doc.width*0.16, doc.width*0.84]
        )
    elif branch_logo:
        header_inner = Table([[branch_logo]], colWidths=[doc.width])
    elif company_logo:
        header_inner = Table([[company_logo]], colWidths=[doc.width])
    elif header_para:
        header_inner = Table([[header_para]], colWidths=[doc.width])
    else:
        header_inner = None

    if header_inner:
        header_inner.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
        ]))
        elements.append(header_inner)
        elements.append(Spacer(1, 0.2*cm))

    # ========== BLUE TITLE BAR ==========
    title_bar = Table(
        [[Paragraph(_prepare_arabic('عرض سعر'), title_style)]],
        colWidths=[doc.width], rowHeights=[1.1*cm]
    )
    title_bar.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), PRIMARY),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
    ]))
    elements.append(title_bar)
    elements.append(Spacer(1, 0.1*cm))
    elements.append(Table([['']], colWidths=[doc.width], rowHeights=[2], style=TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), PRIMARY),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
    ])))
    elements.append(Spacer(1, 0.35*cm))

    # ========== RECIPIENT INFO ==========
    if recipient is None:
        recipient = offer.recipients.select_related('student__contact').first()

    if recipient:
        student = recipient.student
        contact = getattr(student, 'contact', None)
        if student:
            r_name = contact.get_full_name() if contact else student.get_full_name()
            r_phone = contact.mobile if contact else ''
            r_email = contact.email if contact else ''
        else:
            r_name = recipient.contact_name or 'مستلم سريع'
            r_phone = recipient.contact_phone
            r_email = recipient.contact_email

        recipient_rows = [
            [Paragraph(_prepare_arabic(r_name), field_style), Paragraph(_prepare_arabic('اسم المشترك'), section_label_style)],
            [Paragraph(_prepare_arabic(r_phone or '-'), field_style), Paragraph(_prepare_arabic('الهاتف'), section_label_style)],
        ]
        if r_email:
            recipient_rows.append([
                Paragraph(_prepare_arabic(r_email), field_style),
                Paragraph(_prepare_arabic('البريد الإلكتروني'), section_label_style)
            ])
        r_table = Table(recipient_rows, colWidths=[doc.width*0.7, doc.width*0.3])
        r_table.setStyle(TableStyle([
            ('BACKGROUND', (1, 0), (1, -1), PRIMARY_LIGHT),
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 10),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
            ('BOX', (0, 0), (-1, -1), 0.75, PRIMARY),
            ('INNERGRID', (0, 0), (-1, -1), 0.25, BORDER),
        ]))
        elements.append(r_table)
        elements.append(Spacer(1, 0.3*cm))

    # ========== OFFER DETAILS ==========
    course_hours = ''
    if offer.course and offer.course.hours:
        course_hours = f"{offer.course.hours} ساعة"
    elif offer.manual_course_hours:
        course_hours = f"{offer.manual_course_hours} ساعة"
    elif offer.manual_program_hours:
        course_hours = f"{offer.manual_program_hours} ساعة"
    elif offer.master and offer.master.hours:
        course_hours = f"{offer.master.hours} ساعة"

    subscription_display = 'دورة انفرادية' if offer.manual_course_name else offer.title
    detail_rows = [
        [Paragraph(_prepare_arabic(subscription_display), field_style), Paragraph(_prepare_arabic('نوع  الاشتراك'), section_label_style)],
        [Paragraph(_prepare_arabic(str(offer.branch)), field_style), Paragraph(_prepare_arabic('الفرع'), section_label_style)],
        [Paragraph(_prepare_arabic(str(offer.course) if offer.course else (offer.manual_course_name or '-')), field_style), Paragraph(_prepare_arabic('الدورة'), section_label_style)],
        [Paragraph(_prepare_arabic(course_hours or '-'), field_style), Paragraph(_prepare_arabic('عدد الساعات'), section_label_style)],
        [Paragraph(_prepare_arabic(offer.created_at.strftime('%Y-%m-%d')), field_style), Paragraph(_prepare_arabic('تاريخ العرض'), section_label_style)],
        [Paragraph(_prepare_arabic(offer.get_status_display()), field_style), Paragraph(_prepare_arabic('الحالة'), section_label_style)],
    ]
    if offer.start_date:
        detail_rows.append([Paragraph(_prepare_arabic(offer.start_date.strftime('%Y-%m-%d')), field_style), Paragraph(_prepare_arabic('تاريخ البداية'), section_label_style)])
    if offer.end_date:
        detail_rows.append([Paragraph(_prepare_arabic(offer.end_date.strftime('%Y-%m-%d')), field_style), Paragraph(_prepare_arabic('تاريخ النهاية'), section_label_style)])

    dt = Table(detail_rows, colWidths=[doc.width*0.7, doc.width*0.3])
    dt.setStyle(TableStyle([
        ('BACKGROUND', (1, 0), (1, -1), LIGHT_BG),
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 10),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('BOX', (0, 0), (-1, -1), 0.5, BORDER),
        ('INNERGRID', (0, 0), (-1, -1), 0.25, BORDER),
    ]))
    elements.append(dt)
    elements.append(Spacer(1, 0.35*cm))

    # ========== PRICE HIGHLIGHT ==========
    price_label = _prepare_arabic('قيمة العرض')
    currency_name = offer.branch.company.get_currency_display_name() if offer.branch and offer.branch.company else 'ريال'
    price_value = _prepare_arabic(f'{offer.price:,.2f} {currency_name}')
    price_note = _prepare_arabic(offer.price_description) if offer.price_description else ''

    price_inner = [
        [Paragraph(price_label, price_label_center_style)],
        [Paragraph(price_value, price_value_style)],
    ]
    if price_note:
        price_inner.append([Paragraph(price_note, price_note_style)])

    price_box = Table(price_inner, colWidths=[doc.width])
    price_box.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), PRIMARY_LIGHT),
        ('BOX', (0, 0), (-1, -1), 1.5, PRIMARY),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(price_box)
    elements.append(Spacer(1, 0.35*cm))

    # ========== CONTENT BOX ==========
    elements.append(Paragraph(_prepare_arabic('وصف العرض'), section_label_style))
    content_box = Table(
        [[Paragraph(_prepare_arabic_paragraph(offer.content or '-'), content_style)]],
        colWidths=[doc.width]
    )
    content_box.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), LIGHT_BG),
        ('BOX', (0, 0), (-1, -1), 0.5, BORDER),
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('RIGHTPADDING', (0, 0), (-1, -1), 12),
        ('LEFTPADDING', (0, 0), (-1, -1), 12),
    ]))
    elements.append(content_box)
    elements.append(Spacer(1, 0.4*cm))

    # ========== SIGNATURE ==========
    sig_path = branch.signature.path if branch and branch.signature else None
    sig_img = None
    if sig_path:
        try:
            sig_img = Image(sig_path, width=4.0*cm, height=1.8*cm)
        except Exception:
            sig_img = None

    sig_label = Paragraph(_prepare_arabic('التوقيع'), sig_style)
    sig_line_or_img = sig_img or Paragraph(_prepare_arabic('________________________'), sig_style)
    sig_box = Table(
        [[sig_line_or_img], [sig_label]],
        colWidths=[doc.width*0.42]
    )
    sig_box.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
    ]))

    date_box = Paragraph(
        _prepare_arabic(f"التاريخ: {offer.created_at.strftime('%Y-%m-%d')}"),
        sig_style
    )

    sig_row = [
        date_box,
        sig_box,
    ]
    sig_table = Table([sig_row], colWidths=[doc.width*0.5, doc.width*0.5])
    sig_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(sig_table)
    elements.append(Spacer(1, 0.25*cm))

    # ========== FOOTER ==========
    creator = offer.created_by.get_full_name() or offer.created_by.email
    elements.append(Paragraph(_prepare_arabic(f"تم إعداد العرض بواسطة: {creator}"), footer_style))
    if offer.end_date:
        elements.append(Paragraph(_prepare_arabic(f"سريان العرض حتى: {offer.end_date.strftime('%Y-%m-%d')}"), footer_style))

    doc.build(elements)
    buffer.seek(0)
    return buffer


def export_studentoffer_pdf(request, slug):
    """Export offer as a clean, modern one-page PDF."""
    offer = get_object_or_404(StudentOffer, slug=slug)
    if not request.user.has_perm('view_studentoffer', branch=offer.branch):
        raise PermissionDenied('غير مسموح لك دخول هنا')

    recipient_pk = request.GET.get('recipient')
    recipient = None
    if recipient_pk:
        try:
            recipient = offer.recipients.select_related('student__contact').get(pk=recipient_pk)
        except OfferRecipient.DoesNotExist:
            pass
    if not recipient:
        recipient = offer.recipients.select_related('student__contact').first()

    buffer = build_offer_pdf(offer, recipient=recipient)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="offer-{offer.slug}.pdf"'
    return response


# ============================================================
# Bulk Export (filtered list)
# ============================================================

@login_required
def export_offers_excel(request):
    """Export filtered offers list as Excel."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.cell.cell import MergedCell

    user = request.user
    qs = filter_by_branch(
        StudentOffer.objects.all(), user, 'view_studentoffer', perm='view_studentoffer'
    )
    qs = _apply_offer_filters(qs, request.GET)
    qs = qs.select_related('branch', 'course', 'created_by').order_by('-created_at')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'العروض'

    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ['العنوان', 'الفرع', 'الدورة', 'السعر', 'الحالة', 'تاريخ الإنشاء', 'أنشئ بواسطة']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    ws.row_dimensions[1].height = 25

    for row_num, offer in enumerate(qs, 2):
        values = [
            offer.title,
            str(offer.branch),
            str(offer.course) if offer.course else (offer.manual_course_name or '-'),
            float(offer.price),
            offer.status,
            offer.created_at.strftime('%Y-%m-%d %H:%M') if offer.created_at else '',
            offer.created_by.get_full_name() or offer.created_by.email,
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    for col_cells in ws.columns:
        if isinstance(col_cells[0], MergedCell):
            continue
        max_len = 0
        for cell in col_cells:
            try:
                if len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 40)

    now = timezone.localtime().strftime('%Y%m-%d_%H%M')
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="offers_report_{now}.xlsx"'
    wb.save(response)
    return response


@login_required
def export_offers_pdf(request):
    """Export filtered offers list as a PDF report table."""
    import os
    from io import BytesIO
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from django.conf import settings
    from django.contrib.staticfiles.finders import find

    user = request.user
    qs = filter_by_branch(
        StudentOffer.objects.all(), user, 'view_studentoffer', perm='view_studentoffer'
    )
    qs = _apply_offer_filters(qs, request.GET)
    qs = qs.select_related('branch', 'course', 'created_by').order_by('-created_at')

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=1*cm, leftMargin=1*cm,
                            topMargin=1*cm, bottomMargin=1*cm)
    elements = []

    font_name = 'ArialArabic'
    BASE_DIR = settings.BASE_DIR
    candidates = [
        r'C:\Windows\Fonts\arial.ttf',
        r'C:\Windows\Fonts\arialbd.ttf',
        os.path.join(BASE_DIR, 'static', 'fonts', 'Cairo-Regular.ttf'),
        os.path.join(BASE_DIR, 'staticfiles', 'fonts', 'Cairo-Regular.ttf'),
        find('fonts/Cairo-Regular.ttf'),
        find('fonts/Amiri-Regular.ttf'),
    ]
    registered = False
    for font_path in candidates:
        if font_path and os.path.exists(font_path):
            try:
                pdfmetrics.registerFont(TTFont(font_name, font_path))
                from reportlab.pdfbase.pdfmetrics import registerFontFamily
                registerFontFamily(font_name, normal=font_name, bold=font_name, italic=font_name, boldItalic=font_name)
                registered = True
                break
            except Exception:
                continue
    if not registered:
        font_name = 'Helvetica'

    import arabic_reshaper
    from bidi.algorithm import get_display
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT

    def _ar(text):
        if not text:
            return ''
        reshaped = arabic_reshaper.reshape(str(text))
        return get_display(reshaped)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('TitleAr', parent=styles['Normal'], fontName=font_name, fontSize=16, alignment=TA_CENTER, spaceAfter=12)
    header_style = ParagraphStyle('HeaderAr', fontName=font_name, fontSize=9, alignment=TA_CENTER,
                                  textColor=colors.white, leading=14)
    cell_style = ParagraphStyle('CellAr', fontName=font_name, fontSize=8, alignment=TA_CENTER, leading=12)

    elements.append(Paragraph(_ar('تقرير العروض'), title_style))
    elements.append(Spacer(1, 0.3*cm))

    headers = ['العنوان', 'الفرع', 'الدورة', 'السعر', 'الحالة', 'التاريخ', 'أنشئ بواسطة']
    table_data = [[Paragraph(_ar(h), header_style) for h in headers]]

    for offer in qs:
        row = [
            Paragraph(_ar(offer.title), cell_style),
            Paragraph(_ar(str(offer.branch)), cell_style),
            Paragraph(_ar(str(offer.course) if offer.course else (offer.manual_course_name or '-')), cell_style),
            Paragraph(_ar(f'{offer.price}'), cell_style),
            Paragraph(_ar(offer.status), cell_style),
            Paragraph(_ar(offer.created_at.strftime('%Y-%m-%d %H:%M') if offer.created_at else ''), cell_style),
            Paragraph(_ar(offer.created_by.get_full_name() or offer.created_by.email), cell_style),
        ]
        table_data.append(row)

    page_w = landscape(A4)[0] - 2*cm
    col_widths = [page_w*0.22, page_w*0.16, page_w*0.18, page_w*0.10, page_w*0.10, page_w*0.14, page_w*0.10]

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F1F5F9')]),
        ('BOX', (0, 0), (-1, -1), 0.75, colors.HexColor('#2563EB')),
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#CBD5E1')),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 0.5*cm))
    count_style = ParagraphStyle('CountAr', fontName=font_name, fontSize=10, alignment=TA_RIGHT)
    elements.append(Paragraph(_ar(f'إجمالي النتائج: {qs.count()}'), count_style))

    doc.build(elements)

    now = timezone.localtime().strftime('%Y%m-%d_%H%M')
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="offers_report_{now}.pdf"'
    return response


# ============================================================
# Send Offer to Single Recipient
# ============================================================

def _get_recipient_context(offer, recipient):
    """Build context for offer email template."""
    student = recipient.student if recipient else None
    contact = getattr(student, 'contact', None)
    if student:
        recipient_name = contact.get_full_name() if contact else student.get_full_name()
    else:
        recipient_name = recipient.contact_name or 'مستلم سريع' if recipient else 'مستلم سريع'

    branch = offer.branch
    company = branch.company if branch else None
    currency_name = company.get_currency_display_name() if company else 'ريال'

    course_hours = None
    if offer.course and offer.course.hours:
        course_hours = f"{offer.course.hours} ساعة"
    elif offer.manual_course_hours:
        course_hours = f"{offer.manual_course_hours} ساعة"
    elif offer.manual_program_hours:
        course_hours = f"{offer.manual_program_hours} ساعة"
    elif offer.master and offer.master.hours:
        course_hours = f"{offer.master.hours} ساعة"

    return {
        'offer': offer,
        'recipient_name': recipient_name,
        'branch_name': branch.name if branch else '-',
        'course_name': str(offer.course) if offer.course else (offer.manual_course_name or None),
        'master_name': str(offer.master) if offer.master else None,
        'course_hours': course_hours,
        'currency_name': currency_name,
        'creator_name': offer.created_by.get_full_name() or offer.created_by.email,
    }


def send_offer_email(offer, recipient, email):
    """Send offer email with HTML body and PDF attachment."""
    from django.template.loader import render_to_string
    from django.core.mail import EmailMultiAlternatives

    context = _get_recipient_context(offer, recipient)
    subject = offer.title
    html_body = render_to_string('offers/email_offer.html', context)
    text_body = f"""{offer.title}

{offer.content}

السعر: {offer.price} {context['currency_name']}
"""

    pdf_buffer = build_offer_pdf(offer, recipient=recipient)

    msg = EmailMultiAlternatives(
        subject=subject,
        body=text_body,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=[email],
    )
    msg.attach_alternative(html_body, 'text/html')
    msg.attach(
        filename=f"offer-{offer.slug}.pdf",
        content=pdf_buffer.getvalue(),
        mimetype='application/pdf',
    )
    msg.send()


@login_required
def send_offer_to_recipient(request, slug, recipient_pk):
    """Send an offer to a single recipient via their preferred channel."""
    print(""""Send an offer to a single recipient via their preferred channel.""");
    offer = get_object_or_404(StudentOffer, slug=slug)
    if not request.user.has_perm('change_studentoffer', branch=offer.branch):
        raise PermissionDenied('غير مسموح لك دخول هنا')
    recipient = get_object_or_404(OfferRecipient, pk=recipient_pk, offer=offer)

    channel = recipient.channel
    student = recipient.student
    contact = getattr(student, 'contact', None)
    body = f"{offer.title}\n\n{offer.content}"

    # Determine recipient display name
    if student:
        recipient_name = student.get_full_name()
    else:
        recipient_name = recipient.contact_name or recipient.contact_phone or 'مستلم سريع'

    if channel == 'whatsapp':
        phone = contact.mobile if contact else recipient.contact_phone
        if not phone:
            messages.error(request, 'لا يوجد رقم محمول مسجل لهذا المستلم.')
            return redirect('studentoffer-detail', slug=slug)
        try:
            pdf_buffer = build_offer_pdf(offer, recipient)
            result = send_whatsapp_pdf(
                phone,
                filename=f"offer-{offer.slug}.pdf",
                pdf_bytes=pdf_buffer.getvalue(),
                caption=body,
                branch=offer.branch,
            )
            if result.get('success'):
                messages.success(request, f'تم إرسال العرض عبر واتساب إلى {recipient_name}.')
                recipient.status = 'مرسل'
                recipient.sent_at = timezone.now()
                recipient.save()
            else:
                messages.error(request, f'فشل إرسال واتساب: {result.get("error", "خطأ غير معروف")}')
        except Exception as e:
            logger.exception('Failed to send offer WhatsApp to %s', phone)
            messages.error(request, f'فشل إرسال واتساب: {str(e)}')
    elif channel == 'email':
        email = contact.email if contact else recipient.contact_email
        if not email:
            messages.warning(request, 'لا يوجد بريد إلكتروني مسجل لهذا المستلم.')
        else:
            try:
                send_offer_email(offer, recipient, email)
                messages.success(request, f'تم إرسال الإيميل إلى {recipient_name}.')
                recipient.status = 'مرسل'
                recipient.sent_at = timezone.now()
                recipient.save()
            except Exception as e:
                logger.exception('Failed to send offer email to %s', email)
                messages.error(request, f'فشل إرسال الإيميل: {str(e)}')
    elif channel == 'app':
        messages.warning(request, 'قناة إشعار التطبيق غير مفعلة حالياً.')
    else:
        messages.error(request, 'قناة إرسال غير معروفة.')

    return redirect('studentoffer-detail', slug=slug)


@login_required
def add_recipient_to_offer(request, slug):
    """Add a recipient directly to a specific offer (HTML form fallback)."""
    offer = get_object_or_404(StudentOffer, slug=slug)
    if not request.user.has_perm('change_studentoffer', branch=offer.branch):
        raise PermissionDenied('غير مسموح لك دخول هنا')
    if request.method == 'POST':
        form = OfferRecipientAddForm(request.POST, user=request.user)
        if form.is_valid():
            recipient = form.save(commit=False)
            recipient.offer = offer
            try:
                recipient.save()
                messages.success(request, f'تم إضافة المستلم {recipient.student} بنجاح.')
                return redirect('studentoffer-detail', slug=slug)
            except Exception as e:
                messages.error(request, f'خطأ أثناء الحفظ: {str(e)}')
    else:
        form = OfferRecipientAddForm(user=request.user)
    return render(request, 'offers/add_recipient_form.html', {'form': form, 'offer': offer})


def _channel_to_prospect_method(channel):
    mapping = {'whatsapp': 'whatsapp', 'email': 'email'}
    return mapping.get(channel, 'other')


def _create_prospect_from_recipient(offer, recipient, user):
    """Create a Prospect record from a quick/manual offer recipient (Root company only)."""
    prospect = Prospect.objects.create(
        branch=offer.branch,
        name=recipient.contact_name or recipient.contact_phone or 'مستلم سريع',
        mobile=recipient.contact_phone,
        master=offer.master,
        communication_method=_channel_to_prospect_method(recipient.channel),
        status='interested',
        contact_date=timezone.now().date(),
        contacted_by=user if hasattr(user, 'pk') else None,
    )
    return prospect


def _create_prospect_from_root_offer(offer, recipient, user, cd):
    """Create a Prospect record from Root offer form data (with extra fields)."""
    prospect = Prospect.objects.create(
        branch=offer.branch,
        name=recipient.contact_name or recipient.contact_phone or 'مستلم سريع',
        mobile=recipient.contact_phone,
        master=offer.master,
        workplace=cd.get('workplace', ''),
        ministry=cd.get('ministry', ''),
        governorate=cd.get('governorate', ''),
        communication_method=_channel_to_prospect_method(recipient.channel),
        notes=cd.get('notes', ''),
        status='interested',
        contact_date=timezone.now().date(),
        contacted_by=user if hasattr(user, 'pk') else None,
    )
    return prospect


@require_POST
@login_required
def studentoffer_add_recipient_ajax(request, slug):
    """AJAX endpoint to add a recipient to an offer from a modal."""
    offer = get_object_or_404(StudentOffer, slug=slug)
    if not request.user.has_perm('change_studentoffer', branch=offer.branch):
        raise PermissionDenied('غير مسموح لك دخول هنا')
    form = OfferRecipientAddForm(request.POST, user=request.user)
    if form.is_valid():
        recipient = form.save(commit=False)
        recipient.offer = offer
        try:
            recipient.save()
            # في شركة الجذور، المستلم يتحول لمستفسر (مخزن أو جديد)
            if _is_root_branch(offer.branch) and not recipient.student:
                if recipient.prospect:
                    recipient.contact_name = recipient.prospect.name
                    recipient.contact_phone = recipient.prospect.mobile
                    recipient.save(update_fields=['contact_name', 'contact_phone'])
                elif recipient.contact_name or recipient.contact_phone:
                    prospect = _create_prospect_from_recipient(offer, recipient, request.user)
                    recipient.prospect = prospect
                    recipient.save(update_fields=['prospect'])
            if recipient.student:
                name = str(recipient.student)
            elif recipient.prospect:
                name = recipient.prospect.name
            else:
                name = recipient.contact_name or recipient.contact_phone or 'مستلم سريع'
            return JsonResponse({
                'success': True,
                'message': f'تم إضافة المستلم {name} بنجاح.',
                'id': recipient.id,
                'student_name': name,
                'channel': recipient.get_channel_display(),
                'status': recipient.get_status_display(),
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    return JsonResponse({'success': False, 'errors': form.errors}, status=400)


@login_required
def send_offer_to_all(request, slug):
    """Send the offer to all recipients via their preferred channels."""
    offer = get_object_or_404(StudentOffer, slug=slug)
    if not request.user.has_perm('change_studentoffer', branch=offer.branch):
        raise PermissionDenied('غير مسموح لك دخول هنا')
    recipients = offer.recipients.select_related('student__contact').all()
    if not recipients:
        messages.warning(request, 'لا يوجد مستلمون لهذا العرض. أضف مستلمين أولاً.')
        return redirect('studentoffer-detail', slug=slug)

    sent_count = 0
    failed_count = 0
    body = f"{offer.title}\n\n{offer.content}"

    for recipient in recipients:
        channel = recipient.channel
        student = recipient.student
        contact = getattr(student, 'contact', None)

        if channel == 'whatsapp':
            phone = contact.mobile if contact else recipient.contact_phone
            if phone:
                try:
                    pdf_buffer = build_offer_pdf(offer, recipient)
                    result = send_whatsapp_pdf(
                        phone,
                        filename=f"offer-{offer.slug}.pdf",
                        pdf_bytes=pdf_buffer.getvalue(),
                        caption=body,
                        branch=offer.branch,
                    )
                    if result.get('success'):
                        recipient.status = 'مرسل'
                        recipient.sent_at = timezone.now()
                        recipient.save()
                        sent_count += 1
                    else:
                        failed_count += 1
                except Exception as e:
                    logger.exception('Failed to send offer WhatsApp to %s', phone)
                    failed_count += 1
            else:
                failed_count += 1
        elif channel == 'email':
            email = contact.email if contact else recipient.contact_email
            if email:
                try:
                    send_offer_email(offer, recipient, email)
                    recipient.status = 'مرسل'
                    recipient.sent_at = timezone.now()
                    recipient.save()
                    sent_count += 1
                except Exception as e:
                    logger.exception('Failed to send offer email to %s', email)
                    failed_count += 1
            else:
                failed_count += 1
        elif channel == 'app':
            failed_count += 1
        else:
            failed_count += 1

    if sent_count:
        messages.success(request, f'تم الإرسال بنجاح إلى {sent_count} مستلم.')
        offer.status = 'مرسلة'
        offer.sent_at = timezone.now()
        offer.save()
    if failed_count:
        messages.warning(
            request,
            f'فشل الإرسال إلى {failed_count} مستلم. '
            f'تأكد من وجود أرقام واتساب أو بريد إلكتروني مسجل للمستلمين.'
        )

    return redirect('studentoffer-detail', slug=slug)


@login_required
def branch_offer_template_ajax(request, branch_id):
    """AJAX endpoint to fetch the default offer template for a branch."""
    from core.models import Branch
    branch = get_object_or_404(Branch, pk=branch_id)
    if not request.user.is_executive():
        allowed_ids = [b.pk for b in request.user.get_branches_for_perm('add_studentoffer')]
        if branch_id not in allowed_ids:
            raise PermissionDenied('غير مسموح لك دخول هنا')
    return JsonResponse({
        'success': True,
        'template': branch.offer_template or '',
    })


@login_required
def master_courses_ajax(request, master_id):
    """AJAX endpoint to fetch courses for a given master."""
    from courses.models import Master, Course
    master = get_object_or_404(Master, pk=master_id)
    if not request.user.is_executive():
        allowed_ids = [b.pk for b in request.user.get_branches_for_perm('add_studentoffer')]
        if master.branch_id not in allowed_ids:
            raise PermissionDenied('غير مسموح لك دخول هنا')
    courses = Course.objects.filter(master=master).values('id', 'code', 'name', 'master__name').order_by('-created_at')
    return JsonResponse({
        'success': True,
        'courses': list(courses),
    })


@require_POST
@login_required
def root_offer_ajax(request):
    """AJAX endpoint for Root company to create a program/course offer and send it."""
    try:
        if not request.user.has_perm_on_any_branch('add_studentoffer'):
            return JsonResponse({'success': False, 'error': 'غير مسموح لك دخول هنا'}, status=403)

        form = RootQuickOfferForm(request.POST, user=request.user)
        if form.is_valid():
            cd = form.cleaned_data
            master = cd['master']
            branch = cd['branch']

            # Build title and course based on offer type
            if cd['offer_type'] == 'program':
                title = master.name
                course_for_offer = None
                manual_course_name = ''
                manual_course_hours = None
                manual_program_hours = cd.get('program_hours')
                content = cd['content']
            else:
                if not cd.get('course_name'):
                    return JsonResponse({
                        'success': False,
                        'errors': {'course_name': ['يجب كتابة اسم الدورة.']}
                    }, status=400)
                title = cd['course_name']
                course_for_offer = None
                manual_course_name = cd['course_name']
                manual_course_hours = cd.get('course_hours')
                manual_program_hours = None
                renewal_note = 'سيتم تجديد دورات بناء على عدد الساعات والحاجة طبقا لتخصصكم.'
                content = cd['content'].strip()
                if renewal_note not in content:
                    content = (content + '\n\n' + renewal_note).strip()

            offer = StudentOffer.objects.create(
                title=title,
                content=content,
                branch=branch,
                master=master,
                course=course_for_offer,
                manual_course_name=manual_course_name,
                manual_course_hours=manual_course_hours,
                manual_program_hours=manual_program_hours,
                price=cd['price'],
                price_description=cd['price_description'],
                created_by=request.user,
                status='مسودة',
            )

            # في شركة الجذور: إما مستفسر مسجل أو مستفسر جديد
            if cd.get('prospect'):
                prospect = cd['prospect']
                recipient = OfferRecipient.objects.create(
                    offer=offer,
                    student=None,
                    prospect=prospect,
                    contact_name=prospect.name,
                    contact_phone=prospect.mobile,
                    contact_email=cd.get('contact_email', ''),
                    channel=cd['channel'],
                    status='مرسل',
                )
            else:
                recipient = OfferRecipient.objects.create(
                    offer=offer,
                    student=None,
                    contact_name=cd['contact_name'],
                    contact_phone=cd['contact_phone'],
                    contact_email=cd['contact_email'],
                    channel=cd['channel'],
                    status='مرسل',
                )
                # إنشاء مستفسر جديد من بيانات المستلم
                prospect = _create_prospect_from_root_offer(offer, recipient, request.user, cd)
                recipient.prospect = prospect
                recipient.save(update_fields=['prospect'])

            # Send immediately
            channel = cd['channel']
            recipient_name = recipient.contact_name or recipient.contact_phone or 'مستلم سريع'
            body = f"{offer.title}\n\n{offer.content}"
            send_error = None

            if channel == 'whatsapp':
                phone = cd['contact_phone']
                if phone:
                    try:
                        pdf_buffer = build_offer_pdf(offer, recipient)
                        result = send_whatsapp_pdf(
                            phone,
                            filename=f'offer-{offer.slug}.pdf',
                            pdf_bytes=pdf_buffer.getvalue(),
                            caption=body,
                            branch=offer.branch,
                        )
                        if result.get('success'):
                            recipient.status = 'مرسل'
                            recipient.sent_at = timezone.now()
                            recipient.save()
                            offer.status = 'مرسلة'
                            offer.sent_at = timezone.now()
                            offer.save()
                        else:
                            send_error = result.get('error', 'فشل إرسال واتساب')
                    except Exception as e:
                        logger.exception('Failed to send Root offer WhatsApp to %s', phone)
                        send_error = str(e)
                else:
                    send_error = 'لا يوجد رقم محمول مسجل.'
            elif channel == 'email':
                email = cd['contact_email']
                if email:
                    try:
                        send_offer_email(offer, recipient, email)
                        recipient.status = 'مرسل'
                        recipient.sent_at = timezone.now()
                        recipient.save()
                        offer.status = 'مرسلة'
                        offer.sent_at = timezone.now()
                        offer.save()
                    except Exception as e:
                        logger.exception('Failed to send Root offer email to %s', email)
                        send_error = str(e)
                else:
                    send_error = 'لا يوجد بريد إلكتروني مسجل.'
            elif channel == 'app':
                send_error = 'قناة إشعار التطبيق غير مفعلة حالياً.'

            if send_error:
                return JsonResponse({
                    'success': False,
                    'error': send_error,
                    'slug': offer.slug,
                }, status=400)

            return JsonResponse({
                'success': True,
                'message': 'تم إنشاء العرض وإرساله بنجاح.',
                'slug': offer.slug,
            })
        return JsonResponse({'success': False, 'errors': form.errors}, status=400)
    except Exception as exc:
        logger.exception('Error in root_offer_ajax')
        return JsonResponse({'success': False, 'error': str(exc)}, status=400)




# ============================================================
# StudentOffer
# ============================================================

def _apply_offer_filters(queryset, params):
    """Apply shared filter logic for StudentOffer list and export."""
    q = params.get('q')
    if q:
        queryset = queryset.filter(
            Q(title__icontains=q) |
            Q(content__icontains=q) |
            Q(branch__name__icontains=q) |
            Q(course__master__name__icontains=q)
        )

    date_filter = params.get('date_filter')
    if date_filter == 'today':
        queryset = queryset.filter(created_at__date=timezone.localdate())
    elif date_filter == 'this_month':
        today = timezone.localdate()
        queryset = queryset.filter(created_at__year=today.year, created_at__month=today.month)

    date_from = params.get('date_from')
    if date_from:
        queryset = queryset.filter(created_at__date__gte=date_from)

    date_to = params.get('date_to')
    if date_to:
        queryset = queryset.filter(created_at__date__lte=date_to)

    created_by = params.get('created_by')
    if created_by:
        queryset = queryset.filter(created_by_id=created_by)

    interaction = params.get('interaction')
    if interaction == 'accepted':
        queryset = queryset.filter(recipients__status='اشترك').distinct()

    return queryset


class StudentOfferListView(BranchPermissionMixin, ListView):
    model = StudentOffer
    template_name = 'offers/studentoffer_list.html'
    context_object_name = 'offers'
    paginate_by = 25
    required_perm = 'view_studentoffer'
    branch_field = 'branch'

    def get_queryset(self):
        queryset = super().get_queryset()
        queryset = filter_by_branch(queryset, self.request.user, self.branch_field, perm=self.required_perm)
        return _apply_offer_filters(queryset, self.request.GET)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['is_root_user'] = True
        user_branch_ids = [b.pk for b in self.request.user.get_branches_for_perm(self.required_perm)]
        context['all_prospects'] = Prospect.objects.filter(branch__in=user_branch_ids).order_by('-created_at')
        from accounts.models import Person
        offer_user_ids = StudentOffer.objects.filter(
            branch__in=user_branch_ids
        ).values_list('created_by_id', flat=True).distinct()
        context['offer_users'] = Person.objects.filter(pk__in=offer_user_ids).order_by('first_name', 'email')
        context['filter_date'] = self.request.GET.get('date_filter', '')
        context['filter_created_by'] = self.request.GET.get('created_by', '')
        context['filter_interaction'] = self.request.GET.get('interaction', '')
        context['filter_date_from'] = self.request.GET.get('date_from', '')
        context['filter_date_to'] = self.request.GET.get('date_to', '')
        return context


class StudentOfferDetailView(BranchPermissionMixin, DetailView):
    model = StudentOffer
    slug_field = 'slug'
    slug_url_kwarg = 'slug'
    template_name = 'offers/studentoffer_detail.html'
    context_object_name = 'offer'
    required_perm = 'view_studentoffer'
    branch_field = 'branch'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['all_prospects'] = Prospect.objects.filter(branch=self.object.branch).order_by('-created_at')
        return context


class StudentOfferCreateView(BranchPermissionMixin, SuccessMessageMixin, CreateView):
    model = StudentOffer
    form_class = StudentOfferForm
    template_name = 'offers/studentoffer_form.html'
    success_url = reverse_lazy('studentoffer-list')
    success_message = 'تم إنشاء العرض بنجاح.'
    required_perm = 'add_studentoffer'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)


class StudentOfferUpdateView(BranchPermissionMixin, SuccessMessageMixin, UpdateView):
    model = StudentOffer
    slug_field = 'slug'
    slug_url_kwarg = 'slug'
    form_class = StudentOfferForm
    template_name = 'offers/studentoffer_form.html'
    success_url = reverse_lazy('studentoffer-list')
    success_message = 'تم تحديث العرض بنجاح.'
    required_perm = 'change_studentoffer'
    branch_field = 'branch'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs


class StudentOfferDeleteView(BranchPermissionMixin, SuccessMessageMixin, DeleteView):
    model = StudentOffer
    slug_field = 'slug'
    slug_url_kwarg = 'slug'
    template_name = 'offers/studentoffer_confirm_delete.html'
    success_url = reverse_lazy('studentoffer-list')
    success_message = 'تم حذف العرض بنجاح.'
    required_perm = 'delete_studentoffer'
    branch_field = 'branch'


@require_POST
@login_required
def studentoffer_create_ajax(request):
    try:
        if not request.user.has_perm_on_any_branch('add_studentoffer'):
            return JsonResponse({'success': False, 'error': 'غير مسموح لك دخول هنا'}, status=403)
        form = StudentOfferForm(request.POST, user=request.user)
        if form.is_valid():
            offer = form.save(commit=False)
            offer.created_by = request.user
            offer.save()
            return JsonResponse({'success': True, 'message': 'تم إنشاء العرض بنجاح', 'id': offer.id, 'slug': offer.slug})
        return JsonResponse({'success': False, 'errors': form.errors}, status=400)
    except Exception as exc:
        logger.exception('Error in studentoffer_create_ajax')
        return JsonResponse({'success': False, 'error': str(exc)}, status=400)


@require_POST
@login_required
def studentoffer_update_ajax(request, pk):
    try:
        offer = get_object_or_404(StudentOffer, pk=pk)
        if not request.user.has_perm('change_studentoffer', branch=offer.branch):
            return JsonResponse({'success': False, 'error': 'غير مسموح لك دخول هنا'}, status=403)
        form = StudentOfferForm(request.POST, instance=offer, user=request.user)
        if form.is_valid():
            offer = form.save(commit=False)
            offer.save()
            return JsonResponse({'success': True, 'message': 'تم تحديث العرض بنجاح'})
        return JsonResponse({'success': False, 'errors': form.errors}, status=400)
    except Exception as exc:
        logger.exception('Error in studentoffer_update_ajax')
        return JsonResponse({'success': False, 'error': str(exc)}, status=400)


# ============================================================
# OfferRecipient
# ============================================================

class OfferRecipientListView(BranchPermissionMixin, ListView):
    model = OfferRecipient
    template_name = 'offers/offerrecipient_list.html'
    context_object_name = 'recipients'
    paginate_by = 25
    required_perm = 'view_studentoffer'
    branch_field = 'offer__branch'

    def get_queryset(self):
        queryset = super().get_queryset()
        queryset = filter_by_branch(queryset, self.request.user, self.branch_field, perm=self.required_perm)
        q = self.request.GET.get('q')
        if q:
            queryset = queryset.filter(
                Q(offer__title__icontains=q) |
                Q(student__contact__first_name__icontains=q) |
                Q(student__contact__forth_name__icontains=q) |
                Q(contact_name__icontains=q) |
                Q(contact_phone__icontains=q)
            )
        return queryset


class OfferRecipientDetailView(BranchPermissionMixin, DetailView):
    model = OfferRecipient
    template_name = 'offers/offerrecipient_detail.html'
    context_object_name = 'recipient'
    required_perm = 'view_studentoffer'
    branch_field = 'offer__branch'


class OfferRecipientCreateView(BranchPermissionMixin, SuccessMessageMixin, CreateView):
    model = OfferRecipient
    form_class = OfferRecipientForm
    template_name = 'offers/offerrecipient_form.html'
    success_url = reverse_lazy('offerrecipient-list')
    success_message = 'تم إنشاء المستلم بنجاح.'
    required_perm = 'add_studentoffer'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs


class OfferRecipientUpdateView(BranchPermissionMixin, SuccessMessageMixin, UpdateView):
    model = OfferRecipient
    form_class = OfferRecipientForm
    template_name = 'offers/offerrecipient_form.html'
    success_url = reverse_lazy('offerrecipient-list')
    success_message = 'تم تحديث المستلم بنجاح.'
    required_perm = 'change_studentoffer'
    branch_field = 'offer__branch'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs


class OfferRecipientDeleteView(BranchPermissionMixin, SuccessMessageMixin, DeleteView):
    model = OfferRecipient
    template_name = 'offers/offerrecipient_confirm_delete.html'
    success_url = reverse_lazy('offerrecipient-list')
    success_message = 'تم حذف المستلم بنجاح.'
    required_perm = 'delete_studentoffer'
    branch_field = 'offer__branch'


# ============================================================
# OfferNote
# ============================================================

class OfferNoteListView(BranchPermissionMixin, ListView):
    model = OfferNote
    template_name = 'offers/offernote_list.html'
    context_object_name = 'notes'
    paginate_by = 25
    required_perm = 'view_offernote'
    branch_field = 'offer__branch'

    def get_queryset(self):
        queryset = super().get_queryset()
        queryset = filter_by_branch(queryset, self.request.user, self.branch_field, perm=self.required_perm)
        q = self.request.GET.get('q')
        if q:
            queryset = queryset.filter(
                Q(note_text__icontains=q) |
                Q(offer__title__icontains=q)
            )
        return queryset


class OfferNoteDetailView(BranchPermissionMixin, DetailView):
    model = OfferNote
    template_name = 'offers/offernote_detail.html'
    context_object_name = 'note'
    required_perm = 'view_offernote'
    branch_field = 'offer__branch'


class OfferNoteCreateView(BranchPermissionMixin, SuccessMessageMixin, CreateView):
    model = OfferNote
    form_class = OfferNoteForm
    template_name = 'offers/offernote_form.html'
    success_url = reverse_lazy('offernote-list')
    success_message = 'تم إضافة الملاحظة بنجاح.'
    required_perm = 'add_offernote'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        form.instance.person = self.request.user
        return super().form_valid(form)


class OfferNoteUpdateView(BranchPermissionMixin, SuccessMessageMixin, UpdateView):
    model = OfferNote
    form_class = OfferNoteForm
    template_name = 'offers/offernote_form.html'
    success_url = reverse_lazy('offernote-list')
    success_message = 'تم تحديث الملاحظة بنجاح.'
    required_perm = 'change_offernote'
    branch_field = 'offer__branch'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs


class OfferNoteDeleteView(BranchPermissionMixin, SuccessMessageMixin, DeleteView):
    model = OfferNote
    template_name = 'offers/offernote_confirm_delete.html'
    success_url = reverse_lazy('offernote-list')
    success_message = 'تم حذف الملاحظة بنجاح.'
    required_perm = 'delete_offernote'
    branch_field = 'offer__branch'

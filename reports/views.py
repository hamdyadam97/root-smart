from django.contrib.auth.decorators import login_required
from django.contrib.messages.views import SuccessMessageMixin
from django.core.exceptions import PermissionDenied
from django.db.models import Q
from django.urls import reverse_lazy
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.shortcuts import get_object_or_404

from accounts.mixins import BranchPermissionMixin, filter_by_branch
from core.models import Branch
from .models import ReportSnapshot
from .forms import ReportSnapshotForm
from .utils import generate_report_data, get_dashboard_data


def _report_branches(user):
    """Return branches the user may view reports for."""
    if user.is_executive():
        return Branch.objects.all()
    return Branch.objects.filter(pk__in=[b.pk for b in user.get_branches_for_perm('view_report')])


def _can_view_report(user, report):
    """Check whether the user may export/view a specific report snapshot."""
    if user.is_executive():
        return True
    if report.branch:
        return user.has_perm('view_report', report.branch)
    return user.has_perm_on_any_branch('view_report')


def _check_report_perm(user, perm, branch=None):
    """Raise PermissionDenied if the user lacks the report permission (optionally on a branch)."""
    if user.is_executive():
        return
    if branch is not None and not user.has_perm(perm, branch):
        raise PermissionDenied('غير مسموح لك دخول هنا')
    if branch is None and not user.has_perm_on_any_branch(perm):
        raise PermissionDenied('غير مسموح لك دخول هنا')


class ReportSnapshotListView(BranchPermissionMixin, ListView):
    model = ReportSnapshot
    template_name = 'reports/reportsnapshot_list.html'
    context_object_name = 'reports'
    paginate_by = 25
    required_perm = 'view_report'
    branch_field = 'branch'

    def get_queryset(self):
        queryset = super().get_queryset()
        queryset = filter_by_branch(queryset, self.request.user, 'branch')
        q = self.request.GET.get('q')
        if q:
            queryset = queryset.filter(
                Q(report_type__icontains=q) |
                Q(period__icontains=q) |
                Q(branch__name__icontains=q)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['branches'] = _report_branches(self.request.user)
        return context


class ReportSnapshotDetailView(BranchPermissionMixin, DetailView):
    model = ReportSnapshot
    slug_field = 'slug'
    slug_url_kwarg = 'slug'
    template_name = 'reports/reportsnapshot_detail.html'
    context_object_name = 'report'
    required_perm = 'view_report'
    branch_field = None

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['branches'] = _report_branches(self.request.user)
        return context


class ReportSnapshotCreateView(BranchPermissionMixin, SuccessMessageMixin, CreateView):
    model = ReportSnapshot
    form_class = ReportSnapshotForm
    template_name = 'reports/reportsnapshot_form.html'
    success_url = reverse_lazy('reportsnapshot-list')
    success_message = 'تم إنشاء التقرير بنجاح.'
    required_perm = 'add_report'

    def form_valid(self, form):
        _check_report_perm(self.request.user, 'add_report', form.instance.branch)
        form.instance.generated_by = self.request.user
        response = super().form_valid(form)
        # Generate report data automatically
        self.object.data_json = generate_report_data(
            self.object.report_type,
            branch=self.object.branch,
            start_date=self.object.start_date,
            end_date=self.object.end_date,
            user=self.request.user,
        )
        self.object.save(update_fields=['data_json'])
        return response


class ReportSnapshotUpdateView(BranchPermissionMixin, SuccessMessageMixin, UpdateView):
    model = ReportSnapshot
    slug_field = 'slug'
    slug_url_kwarg = 'slug'
    form_class = ReportSnapshotForm
    template_name = 'reports/reportsnapshot_form.html'
    success_url = reverse_lazy('reportsnapshot-list')
    success_message = 'تم تحديث التقرير بنجاح.'
    required_perm = 'change_report'

    def form_valid(self, form):
        _check_report_perm(self.request.user, 'change_report', form.instance.branch)
        response = super().form_valid(form)
        # Regenerate report data automatically
        self.object.data_json = generate_report_data(
            self.object.report_type,
            branch=self.object.branch,
            start_date=self.object.start_date,
            end_date=self.object.end_date,
            user=self.request.user,
        )
        self.object.save(update_fields=['data_json'])
        return response


class ReportSnapshotDeleteView(BranchPermissionMixin, SuccessMessageMixin, DeleteView):
    model = ReportSnapshot
    slug_field = 'slug'
    slug_url_kwarg = 'slug'
    template_name = 'reports/reportsnapshot_confirm_delete.html'
    success_url = reverse_lazy('reportsnapshot-list')
    success_message = 'تم حذف التقرير بنجاح.'
    required_perm = 'delete_report'


@login_required
@require_POST
def reportsnapshot_create_ajax(request):
    form = ReportSnapshotForm(request.POST)
    if form.is_valid():
        report = form.save(commit=False)
        _check_report_perm(request.user, 'add_report', report.branch)
        report.generated_by = request.user
        report.save()
        # Generate report data automatically
        report.data_json = generate_report_data(
            report.report_type,
            branch=report.branch,
            start_date=report.start_date,
            end_date=report.end_date,
            user=request.user,
        )
        report.save(update_fields=['data_json'])
        return JsonResponse({'success': True, 'message': 'تم إنشاء التقرير بنجاح', 'id': report.id, 'slug': report.slug})
    return JsonResponse({'success': False, 'errors': form.errors}, status=400)


@login_required
@require_POST
def reportsnapshot_update_ajax(request, pk):
    report = get_object_or_404(ReportSnapshot, pk=pk)
    _check_report_perm(request.user, 'change_report', report.branch)
    form = ReportSnapshotForm(request.POST, instance=report)
    if form.is_valid():
        report = form.save(commit=False)
        _check_report_perm(request.user, 'change_report', report.branch)
        report.save()
        # Regenerate data
        report.data_json = generate_report_data(
            report.report_type,
            branch=report.branch,
            start_date=report.start_date,
            end_date=report.end_date,
            user=request.user,
        )
        report.save(update_fields=['data_json'])
        return JsonResponse({'success': True, 'message': 'تم تحديث التقرير بنجاح'})
    return JsonResponse({'success': False, 'errors': form.errors}, status=400)


# ============================================================
# Dashboard
# ============================================================

class ReportDashboardView(BranchPermissionMixin, TemplateView):
    template_name = 'reports/dashboard.html'
    required_perm = 'view_report'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        branch_id = self.request.GET.get('branch')
        branch = Branch.objects.filter(pk=branch_id).first() if branch_id else None
        context['dashboard'] = get_dashboard_data(
            user=self.request.user,
            branch=branch
        )
        context['branches'] = _report_branches(self.request.user)
        context['selected_branch'] = branch
        return context


# ============================================================
# Export to Excel
# ============================================================

@login_required
def export_report_excel(request, slug):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    report = get_object_or_404(ReportSnapshot, slug=slug)
    if not _can_view_report(request.user, report):
        raise PermissionDenied('غير مسموح لك دخول هنا')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = report.get_report_type_display()

    # Header
    ws.merge_cells('A1:D1')
    ws['A1'] = report.get_report_type_display()
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    ws['A2'] = 'الفترة:'
    ws['B2'] = report.period or '-'
    ws['A3'] = 'الفرع:'
    ws['B3'] = str(report.branch) if report.branch else 'الكل'
    ws['A4'] = 'تاريخ الإنشاء:'
    ws['B4'] = report.created_at.strftime('%Y-%m-%d')

    row = 6
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    if report.data_json:
        for key, value in report.data_json.items():
            ws.merge_cells(f'A{row}:D{row}')
            cell = ws.cell(row=row, column=1, value=key)
            cell.font = Font(bold=True, size=12)
            cell.fill = PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid')
            row += 1

            if isinstance(value, list) and len(value) > 0 and isinstance(value[0], dict):
                # Table headers
                headers = list(value[0].keys())
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
                    cell.border = thin_border
                row += 1

                for item in value:
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=row, column=col, value=item.get(header, ''))
                        cell.border = thin_border
                    row += 1
            else:
                ws.cell(row=row, column=1, value=str(value))
                row += 1
            row += 1

    # Adjust column widths
    from openpyxl.cell.cell import MergedCell
    for col in ws.columns:
        if isinstance(col[0], MergedCell):
            continue
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{report.slug}.xlsx"'
    wb.save(response)
    return response


# ============================================================
# Export to PDF
# ============================================================

def _prepare_arabic(text):
    """Reshape and reverse Arabic text for ReportLab LTR rendering."""
    if not text:
        return ''
    import arabic_reshaper
    reshaped = arabic_reshaper.reshape(str(text))
    return reshaped[::-1]


def _build_table(data, headers, doc_width, arabic_style):
    """Build a well-formatted ReportLab table with wrapping."""
    from reportlab.platypus import Table, TableStyle, Paragraph
    from reportlab.lib import colors

    # Wrap all cell values in Paragraph for text wrapping
    table_data = []
    # Header row
    header_row = [Paragraph(f'<b>{_prepare_arabic(str(h))}</b>', arabic_style) for h in headers]
    table_data.append(header_row)

    for row in data:
        table_row = []
        for h in headers:
            val = row.get(h, '')
            table_row.append(Paragraph(_prepare_arabic(str(val)), arabic_style))
        table_data.append(table_row)

    # Calculate column widths proportionally based on content length
    num_cols = len(headers)
    base_width = doc_width / num_cols
    col_widths = []
    for i, h in enumerate(headers):
        max_len = len(str(h))
        for row in data:
            max_len = max(max_len, len(str(row.get(h, ''))))
        # Scale width between 0.7x and 1.5x of base width
        factor = min(max(max_len / 15, 0.7), 1.8)
        col_widths.append(base_width * factor)

    # Normalize to fit doc_width
    total = sum(col_widths)
    col_widths = [w / total * doc_width for w in col_widths]

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Cairo'),
        ('FONTNAME', (0, 1), (-1, -1), 'Cairo'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8FAFC')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('WORDWRAP', (0, 0), (-1, -1), True),
    ]))
    return table


@login_required
def export_report_pdf(request, slug):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.units import cm

    report = get_object_or_404(ReportSnapshot, slug=slug)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{report.slug}.pdf"'

    # Detect if report has wide tables -> use landscape
    has_wide_table = False
    if report.data_json:
        for value in report.data_json.values():
            if isinstance(value, list) and len(value) > 0 and isinstance(value[0], dict):
                if len(value[0].keys()) > 5:
                    has_wide_table = True
                    break

    if has_wide_table:
        doc = SimpleDocTemplate(response, pagesize=landscape(A4), rightMargin=1.5*cm, leftMargin=1.5*cm, topMargin=1.5*cm, bottomMargin=1.5*cm)
    else:
        doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)

    elements = []
    styles = getSampleStyleSheet()

    # Register Cairo font for Arabic support
    from reportlab.pdfbase.pdfmetrics import registerFontFamily
    font_path = os.path.join('static', 'fonts', 'Cairo-Regular.ttf')
    if os.path.exists(font_path):
        pdfmetrics.registerFont(TTFont('Cairo', font_path))
        registerFontFamily('Cairo', normal='Cairo', bold='Cairo', italic='Cairo', boldItalic='Cairo')

    arabic_style = ParagraphStyle('Arabic', parent=styles['Normal'], fontName='Cairo', fontSize=9, leading=12, alignment=2)
    title_style = ParagraphStyle('ArabicTitle', parent=styles['Title'], fontName='Cairo', fontSize=18, leading=24, alignment=1)
    heading_style = ParagraphStyle('ArabicHeading', parent=styles['Heading2'], fontName='Cairo', fontSize=14, leading=18, alignment=2)

    # Title
    elements.append(Paragraph(_prepare_arabic(report.get_report_type_display()), title_style))
    elements.append(Spacer(1, 0.5*cm))

    # Info
    info_text = f"الفترة: {report.period or '-'} | الفرع: {str(report.branch) if report.branch else 'الكل'} | تاريخ الإنشاء: {report.created_at.strftime('%Y-%m-%d')}"
    elements.append(Paragraph(_prepare_arabic(info_text), arabic_style))
    elements.append(Spacer(1, 0.5*cm))

    if report.data_json:
        for key, value in report.data_json.items():
            elements.append(Paragraph(f"<b>{_prepare_arabic(key)}</b>", heading_style))
            elements.append(Spacer(1, 0.3*cm))

            if isinstance(value, list) and len(value) > 0 and isinstance(value[0], dict):
                headers = list(value[0].keys())
                table = _build_table(value, headers, doc.width, arabic_style)
                elements.append(table)
            elif isinstance(value, list) and len(value) > 0:
                # Simple list
                for item in value:
                    elements.append(Paragraph(f"• {_prepare_arabic(str(item))}", arabic_style))
            else:
                elements.append(Paragraph(_prepare_arabic(str(value)), arabic_style))
            elements.append(Spacer(1, 0.5*cm))

    doc.build(elements)
    return response
